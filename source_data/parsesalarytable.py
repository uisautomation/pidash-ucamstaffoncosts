#!/usr/bin/env python3
"""
Convert salary spine Excel table to a machine-readable format

Usage:
    parsesalarytable.py (-h | --help)
    parsesalarytable.py [--output=<output>] <input>

Options:

    --output=FILE, -o=FILE      Write output to FILE. If "-", use standard output. [default: -]

"""
import fractions
import itertools
import math
import re
import sys

import dateparser
import docopt
import openpyxl
import yaml

# NOTE: these row and column indices are 1-based :(.

# (minimum, maximum) spine points in table
POINT_VALUE_RANGE = (11, 100)

# starting row and column for data
DATA_START_ROW = 10
DATA_START_COL = 4

# column locations of start of point -> salary mapping tables
SALARY_MAPPING_COLUMNS = [21, 22]

# (row, column) locations of point -> salary mapping table titles
SALARY_MAPPING_TITLES = [
    (DATA_START_ROW-1, col_idx)
    for col_idx in SALARY_MAPPING_COLUMNS
]

# Key for each corresponding grade column
GRADE_KEYS = [
    'T_GRADE', 'GRADE_1', 'GRADE_2', 'GRADE_3', 'GRADE_4', 'GRADE_5', 'GRADE_6', 'GRADE_7',
    'GRADE_8', 'GRADE_9', 'GRADE_10', 'GRADE_11', 'GRADE_12_BAND_1', 'GRADE_12_BAND_2',
    'GRADE_12_BAND_3', 'GRADE_12_BAND_4',
]

# Regular expression which matches actual data (rather than headings, etc).
DATA_VALID_PATTERN = re.compile(r'^(?P<increment>T?[0-9]+)?(?P<contribution>\*)?$')


def main():
    # Parse command line options
    opts = docopt.docopt(__doc__)

    # Load first sheet from workbook
    workbook = openpyxl.load_workbook(opts['<input>'])
    sheet = workbook[workbook.sheetnames[0]]

    # Convert sheet to output table
    output_data = convert(sheet)

    if opts['--output'] == '-':
        yaml.dump(output_data, sys.stdout)
    else:
        with open(opts['--output'], 'w') as fobj:
            yaml.dump(output_data, fobj)


def convert(sheet):
    """
    Convert Excel sheet from HR into data table.

    """

    # An iterator which iterates over point values
    points = range(POINT_VALUE_RANGE[1], POINT_VALUE_RANGE[0]-1, -1)

    # A list of effective dates for each salary mapping column. Note that we require that the cell
    # be of the form "from <date>".
    salary_map_effective_dates = [
        dateparser.parse(
            sheet.cell(row=row_idx, column=col_idx).value.lower().replace('from', '').strip()
        ).date()
        for row_idx, col_idx in SALARY_MAPPING_TITLES
    ]

    # Initialise an empty mapping from point name to salary for each salary mapping column.
    salary_maps = {
        effective_date: {} for effective_date in salary_map_effective_dates
    }

    # Initialise an empty list of salary scales for each grade. Note that, because of the order in
    # which the table is processed this is in reverse order of progression. This is fixed when the
    # final data table is produced.
    grade_scales = {key: [] for key in GRADE_KEYS}

    for point, row_idx in zip(points, itertools.count(DATA_START_ROW)):
        # Note: since point values are monotonic integers it is tempting to perform increments by
        # simply incrementing the point value. This is, strictly speaking, wrong since one must
        # only increment with regard to the salary scale for a particular grade. To guard against
        # this we use string keys for the points which cannot easily be incremented inadvertently.

        # Read the point -> salary mapping for each salary mapping column.
        for date, col_idx in zip(salary_map_effective_dates, SALARY_MAPPING_COLUMNS):
            salary_maps[date][f'POINT_{point}'] = normalise_salary(
                sheet.cell(row=row_idx, column=col_idx).value
            )

        # Look to see if this point is part of any grade.
        for grade_key, col_idx in zip(GRADE_KEYS, itertools.count(DATA_START_COL)):
            # Read the corresponding cell.
            cell = sheet.cell(row=row_idx, column=col_idx)

            # If this cell is empty, this point is not part of the grade.
            if cell.value is None:
                continue

            # Extract the textual representation of the cell
            cell_text = str(cell.value).strip()

            # We need to match the cell value against DATA_VALID_PATTERN since some cells are
            # non-empty but are not part of the grade. These are the grade "labels" which are,
            # annoyingly, placed in the data table itself.
            match = DATA_VALID_PATTERN.match(cell_text)
            if not match:
                continue

            # Record the presence of this point in this grade.
            grade_scales[grade_key].append({
                'point': f'POINT_{point}', 'name': match.group('increment'),
                'isContribution': match.group('contribution') is not None,
            })

    # Return the final data table.
    return {
        'salaries': [
            {
                'effectiveDate': effective_date,
                'mapping': salary_map
            }
            for effective_date, salary_map in salary_maps.items()
        ],
        'grades': [
            {
                'name': grade_key,
                'scale': scale[::-1],  # so that the scale is in increment order
            }
            for grade_key, scale in grade_scales.items()
        ],
    }


def normalise_salary(value):
    """Normalise a cell value into a numeric salary."""
    if value is None:
        return 0
    return _excel_round(value)


def _excel_round(n):
    """
    A version of round() which applies the Excel rule that halves rounds *up* rather than the
    conventional wisdom that they round to the nearest even.

    (The jury is out about whether Excel really rounds away from zero or up but rounding up matches
    the tables produced by HR.)

    """
    # Ensure input is a rational
    n = fractions.Fraction(n)
    if n.denominator == 2:
        # always round up halves
        return math.ceil(n)
    return round(n)


if __name__ == '__main__':
    main()
